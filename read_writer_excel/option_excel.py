from openpyxl import load_workbook,workbook
import os
class Document_path:
    '''文件路径'''
    def read_element_path(self):
        self.parentDirPath = os.path.dirname(os.path.abspath(__file__))
        self.excelPath = self.parentDirPath + u'/element_path.xlsx'
        return self.excelPath
    def read_data_path(self):
        self.parentDirPath1 = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        self.excelPath1 = self.parentDirPath1 + u"/test_case/case.xlsx"
        return self.excelPath1
class Readexcel_element_path:
    '''读取element_path'''
    def __init__(self):
        self.document_path = Document_path()
        self.path = self.document_path.read_element_path()
        self.wb = load_workbook(self.path)
        self.sheet = self.wb['Sheet1']

    def order_number(self,row):
        return self.sheet.cell(row=row,column =1).value
    def test_items(self,row):
        return self.sheet.cell(row=row,column=2).value
    def element_path(self,row):
        ele = self.sheet.cell(row=row,column=3).value

        return ele.split("->")
    def elemen_execute(self,row):
        return self.sheet.cell(row=row, column=4).value


class Readexcel_data(object):
    '''读取case'''
    def __init__(self):
        self.document_path = Document_path()
        self.path = self.document_path.read_data_path()
        self.wb = load_workbook(self.path)
        self.sheet = self.wb['Sheet1']
    def test_order_number(self,row):
        return self.sheet.cell(row, column=1).value
    def test_items(self,row):
        return self.sheet.cell(row, column=2).value  #测试项
    def test_son_items(self,row):
        return self.sheet.cell(row, column=3).value
    def clear_data(self,row):
        return self.sheet.cell(row, column=4).value
    def test_data(self,row):
        data = self.sheet.cell(row,column=5).value  #测试数据
        # print('data',data)
        if data == None:
            return data
        else:
            return data.split("->")
    def test_assert(self,row):
        data = self.sheet.cell(row, column=6).value #断言定位
        return data.split("->")
    def assert_method(self,row):
        data = self.sheet.cell(row,column=7).value #断言方法
        return data.split(",")
    def test_assert_data(self,row):
        data = self.sheet.cell(row,column=8).value #断言数据
        if isinstance(data,int):
            data=str(data)
        return data.split('%')
    def test_except_result(self,row):
        return self.sheet.cell(row,column=9).value #预期结果
if __name__ == "__main__":
    a = Readexcel_element_path()
    b = Readexcel_data()
    print(a.element_path(2))
    print(b.test_data(2))
    print(b.test_assert(2))
    print(b.test_assert_data(2))

