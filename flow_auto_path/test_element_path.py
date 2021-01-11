from assert_data.assert_result import Assert
from error_screenshot.error_web_screenshot import Screenshoot_web
from flow_auto_path.local_element_path import Flow_element
from read_writer_excel.option_excel import *
from read_writer_excel.read_case import Read_case
import time
from log.Log import info
from openpyxl import load_workbook
from test_case_result.test_result_build import Test_result
class test_path:
    def __init__(self, driver):
        self.driver = driver
        self.FLow_element = Flow_element(self.driver)
        self.document_path = Document_path()
        self.read_case =Read_case()
        self.read_case_path = self.read_case.all_file()
        self.element_option_run()
        self.error_pata = ''

    def element_option_run(self):

        '''拿取数据执行'''
        try:
            success = 0
            fail_count = 0
            start_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())
            time_s = time.time()
            table = []
            case_t = 0

            for test_path in self.read_case_path:   #读取case
                case_t+=1
                self.Readexcel_element_path = Readexcel_element_path(test_path)
                self.Readexcel_data = Readexcel_data(test_path)
                self.case_wb = load_workbook(test_path)
                self.case_max = self.case_wb['Sheet1'].max_row
                self.element_max = self.case_wb['Sheet2'].max_row
                for element_row in range(2, self.element_max + 1):
                    test_path_item = self.Readexcel_element_path.test_items(element_row)  # 执行功能项
                    execute = self.Readexcel_element_path.elemen_execute(element_row)  # 是否执行
                    if execute == 'N':
                        continue
                    elif execute == 'Y':
                        for case_row in range(2, self.case_max + 1):
                            test_case_item = self.Readexcel_data.test_items(case_row)  # case功能项
                            path = self.Readexcel_element_path.element_path(element_row)# 拿取定位方法元素操作
                            if test_path_item == test_case_item:
                                table_1 = []
                                data = self.Readexcel_data.test_data(case_row)  # 测试数据
                                test_order_number = self.Readexcel_data.test_order_number(case_row)
                                test_son_data = self.Readexcel_data.test_son_items(case_row)  # 测试子项
                                assert_local_method = self.Readexcel_data.test_assert(case_row)  # 断言定位
                                assert_method = self.Readexcel_data.assert_method(case_row)  # 断言方法
                                assert_data = self.Readexcel_data.test_assert_data(case_row)  # 断言数据
                                except_result = self.Readexcel_data.test_except_result(case_row)  # 预期结果
                                table_1.append(test_path_item+test_son_data)
                                table_1.append(str(case_t)+"_"+str(test_order_number))
                                # print('运行',path)
                                # print("断言",assert_local_method)

                                if data:
                                    clear_data = self.Readexcel_data.clear_data(case_row) # 是否清空输入框
                                if "wb%refresh" not in assert_local_method[0]:
                                    self.opthon_path(path,clear_data,data)
                                if "wb%refresh" in assert_local_method[0]:
                                    time.sleep(0.5)
                                self.assert_data_ = self.opthon_path(assert_local_method, )
                                if self.ele_parameter == 1:
                                    print(2,self.ele_parameter)
                                    table_1.append(self.error_pata)
                                    assert_driver = self.FLow_element.return_driver()
                                    screen = Screenshoot_web(assert_driver)
                                    screen_path = screen.screenshoot()
                                    table_1.append(screen_path.replace("\\", "/"))
                                    table.append(table_1)
                                else:

                                    try:
                                        b = 0
                                        for assert_methodi in assert_method:

                                            Assert(self.assert_data_[0], assert_data[b], assert_methodi)
                                            b += 1
                                        success += 1
                                        table_1.append("pass")
                                        info("%s %s执行成功" % (test_order_number,test_son_data))
                                    except Exception as e:
                                        assert_driver = self.FLow_element.return_driver()
                                        screen = Screenshoot_web(assert_driver)
                                        screen_path = screen.screenshoot()
                                        fail_count += 1
                                        table_1.append("fail")
                                        table_1.append(screen_path.replace("\\", "/"))
                                        info("%s执行失败" % test_son_data)
                                    table.append(table_1)


                if element_row == self.element_max:
                    self.case_wb.close()
        except Exception as e:
            info("运行错误")
        finally:
            time_t = time.time()
            time_time = time_t - time_s
            self.count_time = time.strftime('%M:%S', time.localtime(time_time))
            Test_result(count_time=self.count_time,success=success,fail_count=fail_count,start_time=start_time,table=table)
            self.FLow_element.return_driver().quit()


    def opthon_path(self,path,clear_data=None,data=None,a=0):
        assert_data_ = None
        print(path)
        for i in path:
            self.ele_parameter = 0
            ele_path = i.split('%')
            if ele_path[0] in ['xpath', 'css', 'class', 'id', 'name']:
                try:
                    self.FLow_element.location(ele_path[0], ele_path[1])
                except Exception as e:
                    info("第一次定位执行失败")
                    try:
                        self.FLow_element.location(ele_path[0], ele_path[1])
                    except Exception as e:
                        self.ele_parameter = 1
                        print(1,self.ele_parameter)
                        self.error_pata = "定位执行错误"
                        info("第二次定位执行失败")

                if self.ele_parameter == 1:
                    break
                if ele_path[2] == 'send_keys' and data != None:  # send_keys执行if，其它执行else
                    if a == len(data):
                        a = 0
                    else:
                        try:
                            self.FLow_element.page_operation(send_operation=ele_path[2], data=data[a],
                                                         clear_data=clear_data)
                        except Exception as e:
                            info("第一次操作执行失败")
                            try:
                                self.FLow_element.page_operation(send_operation=ele_path[2], data=data[a],
                                                             clear_data=clear_data)
                            except:
                                self.ele_parameter = 1
                                self.error_pata = "操作执行错误"
                                info("第二次操作执行失败")
                        a += 1
                else:
                    try:
                       assert_data_=self.FLow_element.page_operation(send_operation=ele_path[2])
                    except Exception as e:
                        info("断言数据读取失败")
                        try:
                            assert_data_ = self.FLow_element.page_operation(send_operation=ele_path[2])
                        except Exception as e:
                            self.ele_parameter = 1
                            self.error_pata = "断言执行错误执行错误"
                            info("断言数据二次读取失败")
                if self.ele_parameter == 1:
                    break

            elif ele_path[0] == 'js':
                pass
            elif ele_path[0] == 'wb':
                try:
                    self.FLow_element.web_operation(send_operation=ele_path[1])
                except Exception as e:
                    self.FLow_element.web_operation(send_operation=ele_path[1])
        if assert_data_:
            return assert_data_

